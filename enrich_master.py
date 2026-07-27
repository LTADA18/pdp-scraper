#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
enrich_master.py — สร้าง Master Raw Table ที่เติมข้อมูลแล้ว เป็นไฟล์ใหม่ (เรียงแถวตรงกับของเดิม)
ใช้ **template เดียวกับชีท all_in_one** ของ normalize_pdp.py + คอลัมน์ orig_row / link_status
เติม: product_id, shop_id, url จริง, สถานะลิงก์ (ใช้ได้/เปิดไม่ได้), category,
      และข้อมูลที่ scrape แล้ว (ราคา/sold/หัวข้อ spec) — ที่ยังไม่ scrape = Null

*** ไม่แตะไฟล์ MarketSize เดิม *** (openpyxl ทำ chart/pivot ใน sheet รายงานหาย)
ผลลัพธ์เรียงแถวตรงกับ Master Raw Table -> copy คอลัมน์ไปวางทับได้

ใช้:
    python enrich_master.py --master "…/MarketSize-Share (1).xlsx" ^
        --ids "…/MarketSize-Share (1)_ids.xlsx" --raw "data/raw_*.ndjson"
"""
import argparse
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path

import openpyxl

import normalize_pdp as N          # ใช้ build_rows/spec_matrix/combined + template เดียวกัน

TH = timezone(timedelta(hours=7))


def link_status(url):
    u = url or ""
    if not u.strip():
        return "เปิดไม่ได้ (ไม่มีลิงก์)"
    if "/search" in u or "keyword=" in u:
        return "เปิดไม่ได้ (ลิงก์ค้นหา)"
    if re.search(r"s\.lazada\.co\.th|vt\.tiktok\.com", u):
        return "ใช้ได้ (ลิงก์สั้น)"
    if re.search(r"-i\d{6,}|/product/\d+/\d+|i\.\d+\.\d+|/pdp/[^/]+/\d+|/view/product/\d+", u):
        return "ใช้ได้"
    return "ไม่แน่ใจ (ลิงก์รูปแบบอื่น)"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--master", required=True)
    ap.add_argument("--ids", required=True)
    ap.add_argument("--raw", nargs="+", default=["data/raw_*.ndjson"])
    ap.add_argument("--spec-map", default="spec_map.csv")
    ap.add_argument("--category-map", default="category_map.csv")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    run_date = datetime.now(TH).strftime("%Y-%m-%d")
    spec_map = N.load_spec_map(args.spec_map)
    cat_map = N.load_category_map(args.category_map)

    # ---- ข้อมูลที่ scrape แล้ว -> ตาราง all_in_one (template เดียวกับ normalize) ----
    import glob
    paths = [p for g in args.raw for p in glob.glob(g)]
    recs = N.load_records(paths) if paths else []
    prod, var, spec, _ = N.build_rows(recs, run_date, spec_map, cat_map)
    combined = N.build_combined(prod, N.build_spec_matrix(spec, prod))
    template = list(combined.columns) if len(combined) else list(N.PRODUCT_COLS)
    # pid -> แถว all_in_one (dict) ; key = product_id string
    by_pid = {}
    for _, row in combined.iterrows() if len(combined) else []:
        by_pid[str(row["product_id"])] = row.to_dict()

    # ---- master: product_name (B) + category (F) เรียงตามแถว ----
    mwb = openpyxl.load_workbook(args.master, read_only=True)
    mws = mwb["Master Raw Table"]
    mhdr = [c.value for c in next(mws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(mhdr) if n}
    m_name, m_cat = {}, {}
    for ri, row in enumerate(mws.iter_rows(min_row=2, values_only=True), start=2):
        m_name[ri] = row[ci["product_name"]] if "product_name" in ci and ci["product_name"] < len(row) else None
        m_cat[ri] = row[ci["category"]] if "category" in ci and ci["category"] < len(row) else None

    # ---- _ids: orig_row, platform, category, url, product_id, shop_id ----
    iwb = openpyxl.load_workbook(args.ids, read_only=True)
    iws = iwb.active
    ids_rows = list(iws.iter_rows(min_row=2, values_only=True))

    out_cols = ["orig_row", "link_status"] + template
    out = openpyxl.Workbook()
    o = out.active
    o.title = "Master enriched"
    o.append(out_cols)

    NULL = "Null"
    n_id = n_usable = n_scraped = 0
    for r in ids_rows:
        orig, platform, category, url, pid, shop = r[0], r[1], r[2], r[3], r[4], r[5]
        st = link_status(url)
        if st.startswith("ใช้ได้"):
            n_usable += 1
        if pid:
            n_id += 1
        base = by_pid.get(str(pid)) if pid else None      # ข้อมูล scrape (ถ้ามี)
        if base:
            n_scraped += 1

        rowvals = {c: NULL for c in template}
        if base:
            for c in template:
                v = base.get(c)
                rowvals[c] = NULL if v is None or (isinstance(v, float) and v != v) or v == "" else v
        # ค่าที่รู้จาก master/_ids เสมอ (ทับของ scrape ถ้าจำเป็น)
        rowvals["platform"] = platform or rowvals.get("platform") or NULL
        rowvals["category"] = m_cat.get(orig) or category or rowvals.get("category") or NULL
        rowvals["product_id"] = pid or rowvals.get("product_id") or NULL
        rowvals["shop_id"] = shop or rowvals.get("shop_id") or NULL
        rowvals["url"] = url or rowvals.get("url") or NULL
        if rowvals.get("product_name") in (None, NULL, ""):
            rowvals["product_name"] = m_name.get(orig) or NULL

        o.append([orig, st] + [rowvals.get(c, NULL) for c in template])

    dst = Path(args.out) if args.out else Path(args.master).with_name(
        Path(args.master).stem + "_enriched.xlsx")
    out.save(dst)
    print(f"แถว: {len(ids_rows)} | มี product_id: {n_id} | ลิงก์ใช้ได้: {n_usable} | เติมจาก scrape: {n_scraped}")
    print(f"คอลัมน์: {len(out_cols)} (template เดียวกับ all_in_one)")
    print(f"ไฟล์: {dst}")


if __name__ == "__main__":
    main()
