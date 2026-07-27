#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_maps.py — สร้างตาราง mapping 2 ไฟล์ (แก้เองได้ทีหลัง ไม่ต้องแตะโค้ด)
  1) spec_map.csv     : ชื่อ spec ดิบของแต่ละแพลตฟอร์ม -> หัวข้อกลาง (canonical)
  2) category_map.csv : product_id -> category (จากไฟล์ MarketSize-Share _ids.xlsx)

ใช้:
    python build_maps.py --raw data/raw_*.ndjson --ids "C:/.../MarketSize-Share (1)_ids.xlsx"
normalize_pdp.py จะอ่าน 2 ไฟล์นี้ไปใช้ตอนแปลง Excel
"""
import argparse
import csv
import glob
import json
from collections import Counter
from pathlib import Path

# ---- canonical taxonomy: ชื่อ spec ดิบ (ทุกภาษา) -> หัวข้อกลาง ----
# ความหมายตรงกันแต่ชื่อต่างกันข้ามแพลตฟอร์ม รวมเป็นหัวข้อเดียว
CANON = {
    "แบรนด์ (Brand)": ["Brand", "แบรนด์"],
    "รุ่น (Model)": ["Model", "รุ่น", "หมายเลขรุ่น"],
    # ประเภทเครื่องมือ: รวม "ประเภท X" ของทุกหมวดเป็นหัวข้อเดียว (category บอกหมวดอยู่แล้ว)
    "ประเภทเครื่องมือ (Tool Type)": [
        "Drill Type", "ประเภทสว่าน", "ประเภทของสว่าน",
        "Grinder Type", "Welder Type", "Saw Type", "Air Compressor Type", "Outdoor Power Tool Type",
        "Extractor Type", "Pruner Type", "Hand Tools Type", "Tool_Type", "Tools type", "Type Of Mowers",
        "Type Of Planers", "Type Of Demolition Hammers", "Type Of Rotary Tools", "Type Of Sanders",
        "Type Of Nail Guns", "Type Of Heat Guns", "Type Of Chisel", "Type Grease Guns"],
    "ไร้สาย (Cordless)": ["Cordless", "ประเภทไร้สาย", "ไร้สาย"],
    "แรงดันไฟฟ้า (Input Voltage)": ["Input Voltage", "แรงดันไฟฟ้านำเข้า", "แรงดันไฟฟ้าขาเข้า (V)", "แรงดันไฟฟ้า"],
    "กำลังไฟ (Power)": ["Power Consumption", "กำลัง", "กำลังไฟ (W)", "กำลังไฟ", "Power Consumption (W)", "Wattage"],
    "พื้นผิวที่ใช้ได้ (Surface Compatibility)": [
        "Tool Compatible With Surfaces", "พื้นผิวของเครื่องมือ",
        "ความเข้ากันได้ของพื้นผิวเครื่องมือ", "ใช้ได้กับพื้นผิวอะไร"],
    "คุณสมบัติเครื่องมือ (Power Tool Feature)": [
        "Power Tool Feature", "คุณสมบัติของสินค้า", "คุณสมบัติเครื่องมือไฟฟ้า", "Hand Tool Features"],
    "ประเภทแบตเตอรี่ (Battery Core Type)": ["Battery Core Type", "ประเภทแกนแบตเตอรี่", "Battery_Type"],
    "จำนวนแบตเตอรี่ (No. of Batteries)": ["No. of included Batteries"],
    "ประเภทการรับประกัน (Warranty Type)": [
        "Warranty Type", "Warranty", "warranty", "ประเภทการประกัน", "ประเภทของการรับประกัน", "ประเภทการรับประกัน"],
    "ระยะเวลารับประกัน (Warranty Period)": [
        "Warranty Period", "ระยะเวลาการรับประกัน", "ระยะเวลารับประกัน"],
    "วัสดุ (Material)": ["Material", "วัสดุที่ใช้ในการผลิต"],
    "ประเภทสินค้า (Product Type)": ["ประเภทของสินค้า", "Product Feature"],
    "อุปกรณ์ไฟฟ้า (Electric)": ["อุปกรณ์ไฟฟ้า"],
    "แหล่งผลิต (Place of Origin)": ["Place of Origin", "Place Of Origin", "Country of Origin"],
    "มอก. (TIS Certificate)": ["หมายเลขมอก.", "TIS Certificate No."],
    "ขนาด (Dimensions)": ["ขนาด (ยาว x กว้าง x สูง)"],
    "ส่งจาก (Ship From)": ["ส่งจาก"],
    "ระดับทักษะ (Skill Level)": ["ระดับของทักษะ"],
    "หมวดหมู่ Shopee (Category)": ["Category"],
    "ปลั๊ก (Plug Type)": ["Plug Type"],
    "อุปกรณ์ที่ให้มา (Tools Included)": ["Tools Included", "Tool Type Of Accessories"],
}
RAW2CANON = {raw: canon for canon, raws in CANON.items() for raw in raws}


def build_spec_map(raw_globs, out):
    counts = Counter()
    for g in raw_globs:
        for fn in glob.glob(g):
            for line in Path(fn).read_text(encoding="utf-8").splitlines():
                line = line.strip()
                if not line:
                    continue
                try:
                    r = json.loads(line)
                except json.JSONDecodeError:
                    continue
                for s in (r.get("spec") or []):
                    n = (s.get("name") or "").strip()
                    if n:
                        counts[n] += 1

    rows = []
    for raw, c in counts.most_common():
        canon = RAW2CANON.get(raw, "")          # ว่าง = ยังไม่ได้ map (คุณเติมเองได้)
        rows.append((raw, canon, c))
    with open(out, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["raw_spec_name", "canonical", "count"])
        w.writerows(rows)
    mapped = sum(1 for _, c, _ in rows if c)
    print(f"spec_map.csv: {len(rows)} ชื่อ spec ({mapped} map แล้ว, {len(rows)-mapped} ยังว่างให้เติมเอง)")


def build_category_map(ids_path, out):
    import openpyxl
    wb = openpyxl.load_workbook(ids_path, read_only=True)
    ws = wb.active
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {n: i for i, n in enumerate(hdr) if n}
    seen = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        pid = r[ci["product_id"]] if "product_id" in ci else None
        cat = r[ci["category"]] if "category" in ci else None
        if pid and cat and str(pid) not in seen:
            seen[str(pid)] = str(cat)
    with open(out, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["product_id", "category"])
        for pid, cat in seen.items():
            w.writerow([pid, cat])
    print(f"category_map.csv: {len(seen)} product_id -> category")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--raw", nargs="+", default=["data/raw_*.ndjson"])
    ap.add_argument("--ids", help="ไฟล์ *_ids.xlsx จาก fill_ids.py (สำหรับ category)")
    args = ap.parse_args()
    build_spec_map(args.raw, "spec_map.csv")
    if args.ids:
        build_category_map(args.ids, "category_map.csv")


if __name__ == "__main__":
    main()
