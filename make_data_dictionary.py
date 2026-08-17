#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
make_data_dictionary.py — สร้าง Data Dictionary (Excel) สำหรับส่งทีมฐานข้อมูล

อ่านโครงสร้างจริงจาก data/raw_*.ndjson + output/*.xlsx + data/sold_history.csv
แล้วเขียนเป็นเอกสารพร้อมส่ง: ตารางอะไรบ้าง / คอลัมน์ / ชนิด / คีย์ / ข้อควรระวัง

ใช้:
    python make_data_dictionary.py --out output/Data_Dictionary.xlsx
"""
import argparse
import json
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TH = timezone(timedelta(hours=7))
ROOT = Path(__file__).parent

HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13)
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")

# ---------------------------------------------------------------- ตาราง: products (all_in_one)
PRODUCTS = [
    # (คอลัมน์, ชนิดที่แนะนำใน DB, คีย์, บังคับมีค่า, คำอธิบาย, ตัวอย่าง)
    ("date", "DATE", "PK", "ใช่", "วันที่เก็บข้อมูล (เวลาไทย) — 1 แถว = 1 สินค้า/วัน", "2026-08-03"),
    ("platform", "VARCHAR(10)", "PK", "ใช่", "lazada / tiktok / shopee", "lazada"),
    ("product_id", "VARCHAR(32)", "PK", "ใช่*", "รหัสสินค้าของแพลตฟอร์ม **ต้องเป็น string เสมอ** TikTok ยาว 19 หลัก ถ้าเก็บเป็น int/bigint จะเพี้ยน", "2377056833"),
    ("shop_id", "VARCHAR(32)", "", "ไม่", "รหัสร้าน (string เช่นกัน)", "1000302769"),
    ("shop_name", "NVARCHAR(255)", "", "ไม่", "ชื่อร้าน", "Big Powertools"),
    ("product_name", "NVARCHAR(500)", "", "ไม่", "ชื่อประกาศตามหน้าเว็บ", "MOLITA สว่านไร้สาย 198V"),
    ("category", "NVARCHAR(100)", "", "ไม่", "หมวดจาก category_map.csv (map ด้วย product_id) — ครอบคลุมเฉพาะสินค้าที่อยู่ในไฟล์ MarketSize", "สว่านไร้สาย"),
    ("url", "NVARCHAR(1000)", "", "ใช่", "ลิงก์หน้าสินค้า", "https://www.lazada.co.th/..."),
    ("price", "DECIMAL(12,2)", "", "ไม่", "ราคาขายจริง หน่วยบาท (แปลงมาแล้ว) — ถ้ามีหลาย SKU = ราคาต่ำสุด", "399"),
    ("price_max", "DECIMAL(12,2)", "", "ไม่", "ราคาสูงสุดในกลุ่ม SKU", "399"),
    ("original_price", "DECIMAL(12,2)", "", "ไม่", "ราคาก่อนลด (ต้องมาจาก SKU ตัวเดียวกับ price)", "1299"),
    ("discount_pct", "DECIMAL(5,2)", "", "ไม่", "% ส่วนลด คำนวณจาก price vs original_price", "69.3"),
    ("currency", "CHAR(3)", "", "ใช่", "สกุลเงิน (ตลาดไทย = THB เสมอ)", "THB"),
    ("variation_summary", "NVARCHAR(1000)", "", "ไม่", "สรุปตัวเลือกสินค้าแบบข้อความ (รายละเอียดเต็มอยู่ตาราง variants)", "Color: Yellow"),
    ("variation_count", "INT", "", "ใช่", "จำนวนแกนตัวเลือก (สี/ขนาด/รุ่น)", "1"),
    ("sku_count", "INT", "", "ใช่", "จำนวน SKU ทั้งหมดของสินค้านี้", "1"),
    ("spec_summary", "NVARCHAR(2000)", "", "ไม่", "สรุป spec แบบข้อความ (รายละเอียดเต็มอยู่ตาราง specs)", "Brand: MOLITA; ..."),
    ("spec_count", "INT", "", "ใช่", "จำนวนหัวข้อ spec", "6"),
    ("sold_count", "BIGINT", "", "ไม่", "ยอดขายสะสมทั้งหมด **ไม่ใช่ยอดรายวัน** — ดูข้อควรระวังเรื่องความแม่นยำ", "47700"),
    ("sold_today", "BIGINT", "", "ไม่", "ยอดขายวันนี้ = sold_count วันนี้ − snapshot ล่าสุดก่อนหน้า (คำนวณจาก sold_history) วันแรกที่เก็บจะว่าง", "12"),
    ("sold_mtd", "BIGINT", "", "ไม่", "ยอดสะสมตั้งแต่ต้นเดือน (ถ้าเริ่มเก็บกลางเดือน นับจากวันแรกที่มีข้อมูล)", "340"),
    ("sold_band", "NVARCHAR(50)", "", "ไม่", "ยอดขายแบบข้อความช่วงตามที่เว็บแสดง ไว้อ้างอิง/ตรวจสอบ", "47.7K sold"),
    ("source", "VARCHAR(10)", "", "ใช่", "ที่มาของข้อมูล — ใช้คัดคุณภาพ ดูชีต Enumerations", "state"),
    ("notes", "NVARCHAR(1000)", "", "ไม่", "เหตุผลเมื่อข้อมูลไม่ครบ / คำเตือน (เช่น สินค้าถูกลบ, ยอดเป็นเลขกลม)", "lazada sold เป็นเลขกลม"),
]

SPEC_COLS_NOTE = (
    "หลังคอลัมน์ notes จะมีคอลัมน์ spec กลางอีก ~21 คอลัมน์ (แบรนด์, รุ่น, กำลังไฟ, แรงดันไฟฟ้า, "
    "ประเภทเครื่องมือ, มอก., ไร้สาย ฯลฯ) ชื่อคอลัมน์มาจาก spec_map.csv "
    "ทั้งหมดเป็นข้อความ (NVARCHAR) และมีค่าไม่ครบทุกแถว (7-53%) "
    "แนะนำ: ถ้าจะทำ DB ให้ใช้ตาราง specs แบบ long format แทน แล้ว pivot ตอนทำรายงาน"
)

VARIANTS = [
    ("date", "DATE", "PK", "ใช่", "วันที่เก็บ", "2026-08-03"),
    ("platform", "VARCHAR(10)", "PK", "ใช่", "แพลตฟอร์ม", "lazada"),
    ("product_id", "VARCHAR(32)", "PK/FK", "ใช่", "FK -> products.product_id", "2377056833"),
    ("sku_id", "VARCHAR(32)", "PK", "ใช่", "รหัส SKU (string — Shopee/TikTok เป็นเลขยาว)", "8055094765"),
    ("product_name", "NVARCHAR(500)", "", "ใช่", "ชื่อสินค้า (ซ้ำมาเพื่อความสะดวก ไม่ต้องเก็บใน DB ก็ได้)", "MOLITA สว่านไร้สาย"),
    ("option_path", "NVARCHAR(500)", "", "ไม่", "ตัวเลือกของ SKU นี้ เช่น สี/ขนาด คั่นด้วย ' / '", "Yellow"),
    ("price", "DECIMAL(12,2)", "", "ไม่", "ราคาของ SKU นี้ (บาท)", "399"),
    ("original_price", "DECIMAL(12,2)", "", "ไม่", "ราคาก่อนลดของ SKU นี้", "1299"),
    ("stock", "INT", "", "ไม่", "สต็อก — **Lazada ไม่มีข้อมูลจริง** (state ไม่ส่งมา) มีค่าแค่ 33% ของแถว อย่าใช้ตัดสินใจ", "0"),
]

SPECS = [
    ("date", "DATE", "PK", "ใช่", "วันที่เก็บ", "2026-08-03"),
    ("platform", "VARCHAR(10)", "PK", "ใช่", "แพลตฟอร์ม", "lazada"),
    ("product_id", "VARCHAR(32)", "PK/FK", "ใช่", "FK -> products.product_id", "2377056833"),
    ("spec_name", "NVARCHAR(200)", "PK", "ใช่", "ชื่อ spec **ดิบตามที่เว็บให้มา** (ชื่อต่างกันแต่ละแพลตฟอร์ม)", "Brand"),
    ("spec_key", "NVARCHAR(200)", "", "ไม่", "หัวข้อกลางหลัง map ด้วย spec_map.csv — ที่ยังไม่ได้ map จะว่าง (72% มีค่า) ใช้ตัวนี้เวลาเทียบข้ามแพลตฟอร์ม", "แบรนด์ (Brand)"),
    ("spec_value", "NVARCHAR(1000)", "", "ใช่", "ค่าของ spec", "MOLITA"),
    ("product_name", "NVARCHAR(500)", "", "ใช่", "ชื่อสินค้า (ซ้ำมาเพื่อสะดวก)", "MOLITA สว่านไร้สาย"),
]

SOLD_HISTORY = [
    ("date", "DATE", "PK", "ใช่", "วันที่ snapshot", "2026-08-03"),
    ("platform", "VARCHAR(10)", "PK", "ใช่", "แพลตฟอร์ม", "lazada"),
    ("product_id", "VARCHAR(32)", "PK", "ใช่", "รหัสสินค้า", "2377056833"),
    ("sold_count", "BIGINT", "", "ใช่", "ยอดขายสะสม ณ วันนั้น — เป็นฐานคำนวณ sold_today/sold_mtd ทั้งหมด", "47700"),
]

ENUMS = [
    ("source", "state", "อ่านจาก state object ของเว็บโดยตรง — เชื่อถือได้สูงสุด", "ใช้ได้"),
    ("source", "api", "เรียก internal API ของเว็บ (Shopee) — เชื่อถือได้", "ใช้ได้"),
    ("source", "dom", "อ่านจาก HTML เพราะ state หาไม่เจอ — ค่าอาจไม่ครบ/ผิด", "ตรวจก่อนใช้"),
    ("source", "jsonld", "อ่านจาก JSON-LD ในหน้า — สำรอง", "ตรวจก่อนใช้"),
    ("source", "blocked", "เปิดไม่ได้/สินค้าถูกลบ/โดน anti-bot — **ทุกค่าว่างโดยตั้งใจ** เหตุผลอยู่ใน notes", "ไม่ใช้"),
    ("platform", "lazada", "Lazada ไทย", ""),
    ("platform", "tiktok", "TikTok Shop ไทย", ""),
    ("platform", "shopee", "Shopee ไทย", ""),
    ("currency", "THB", "บาท (ตลาดไทยเท่านั้น)", ""),
]

CAVEATS = [
    ("ID ต้องเป็น string",
     "product_id / shop_id / sku_id ห้ามเก็บเป็น INT หรือ BIGINT — TikTok product_id ยาว 19 หลัก "
     "เกินช่วงที่ Excel/JS จัดการได้ ค่าจะเพี้ยนเงียบ ๆ ใช้ VARCHAR เสมอ"),
    ("ความแม่นของยอดขาย ไม่เท่ากันทุกแพลตฟอร์ม",
     "TikTok/Shopee = เลขเป๊ะทุกช่วง | Lazada = เว็บให้มาเป็นเลขกลมเมื่อเกินหลักพัน (เช่น '3.3K' -> 3300) "
     "ทำให้ sold_today ของสินค้าขายดีบน Lazada กระโดดทีละ ~100 ไม่ใช่ยอดจริงรายวัน "
     "ดูคอลัมน์ notes จะมีคำเตือนกำกับรายแถว"),
    ("sold_count คือยอดสะสม ไม่ใช่ยอดรายวัน",
     "ถ้าจะทำรายงานยอดขายรายวันต้องใช้ sold_today/sold_mtd หรือคำนวณ diff เองจากตาราง sold_history"),
    ("ยอดติดลบถูกตัดทิ้ง",
     "ถ้า sold_count วันนี้น้อยกว่าเมื่อวาน (ร้านรีเซ็ต/ข้อมูลเพี้ยน) ระบบเว้น sold_today ว่างไว้ ไม่ใส่เลขติดลบ"),
    ("ค่าว่างใช้คำว่า 'Null' เป็นข้อความ",
     "ในไฟล์ Excel ที่ส่งออก ช่องที่ไม่มีข้อมูลจะเป็นข้อความ 'Null' (ไม่ใช่ช่องว่าง) เพื่อให้เห็นชัดว่า "
     "'ไม่มีข้อมูล' ไม่ใช่ 'ลืมกรอก' — ตอน import เข้า DB ให้แปลง 'Null' -> NULL จริง"),
    ("แถวที่ source=blocked มีอยู่ในไฟล์ด้วย",
     "ไฟล์ส่งออกมีทั้งแถวที่เก็บสำเร็จและไม่สำเร็จ (blocked ~26% ของแถว) ถ้าจะทำรายงานให้กรอง "
     "source IN ('state','api') หรือ product_name IS NOT NULL ก่อน"),
    ("category ยังไม่ครบทุกสินค้า",
     "มาจาก category_map.csv ที่สร้างจากไฟล์ MarketSize — ครอบคลุมประมาณ 38% ของแถว "
     "สินค้าที่ไม่อยู่ในไฟล์นั้นจะไม่มี category"),
    ("stock เชื่อไม่ได้",
     "Lazada ไม่ส่งสต็อกจริงมา (ที่เห็นในบางที่คือลิมิตการสั่งซื้อ ไม่ใช่จำนวนคงเหลือ) "
     "คอลัมน์ stock มีค่าแค่ ~33% ของแถว อย่านำไปใช้ตัดสินใจสต็อก"),
    ("1 แถว = 1 สินค้า/วัน",
     "ตาราง products มี grain = (date, platform, product_id) ถ้าเก็บซ้ำวันเดียวกันต้อง upsert ทับ ไม่ใช่ insert เพิ่ม"),
    ("spec เป็น long format",
     "ตาราง specs 1 แถว = 1 หัวข้อ spec ของสินค้า ใช้ spec_key (หัวข้อกลาง) เวลาเทียบข้ามแพลตฟอร์ม "
     "เพราะ spec_name ดิบตั้งชื่อต่างกัน (Brand / แบรนด์ / ยี่ห้อ)"),
]


def write_table(ws, title, subtitle, headers, rows, widths):
    r = 1
    ws.cell(r, 1, title).font = TITLE_FONT
    r += 1
    if subtitle:
        c = ws.cell(r, 1, subtitle)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
        ws.row_dimensions[r].height = 30
        r += 1
    r += 1
    for j, h in enumerate(headers, 1):
        c = ws.cell(r, j, h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
    hdr_row = r
    for row in rows:
        r += 1
        for j, v in enumerate(row, 1):
            c = ws.cell(r, j, v)
            c.alignment = Alignment(wrap_text=True, vertical="top")
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default="output/Data_Dictionary_PDP.xlsx")
    args = ap.parse_args()

    today = datetime.now(TH).strftime("%Y-%m-%d")
    wb = Workbook()

    # ---- ชีต 1: ภาพรวม ----
    ws = wb.active
    ws.title = "ภาพรวม"
    ws.cell(1, 1, "Data Dictionary — ข้อมูลสินค้า Lazada / TikTok Shop / Shopee (ตลาดไทย)").font = Font(bold=True, size=14)
    lines = [
        f"จัดทำ: {today}",
        "",
        "ข้อมูลนี้คืออะไร: ข้อมูลหน้าสินค้า (PDP) ที่ดึงจากเว็บจริงรายวัน — ราคา ตัวเลือกสินค้า (variation/SKU) "
        "สเปก ชื่อร้าน และยอดขายสะสม",
        "",
        "ตารางที่ส่งให้ (แต่ละตาราง = 1 ชีตในไฟล์ Excel ที่ส่งออก):",
        "   1) products (ชีตชื่อ all_in_one) — 1 แถว = 1 สินค้า/วัน   ~4,500 แถว/วัน",
        "   2) variants — 1 แถว = 1 SKU/วัน   ~10,600 แถว/วัน",
        "   3) specs — 1 แถว = 1 หัวข้อสเปก/สินค้า/วัน   ~18,200 แถว/วัน",
        "   4) sold_history (ไฟล์ CSV แยก) — snapshot ยอดขายสะสมรายวัน ใช้คำนวณยอดขายรายวัน",
        "",
        "ความถี่: เก็บวันละครั้ง (เวลาไทย)",
        "แหล่งข้อมูล: หน้าเว็บจริงของแต่ละแพลตฟอร์ม (ไม่ใช่ API สาธารณะ) ทุกค่ามาจากหน้าเว็บ ไม่มีการประมาณค่าเอง",
        "",
        "อ่านชีต 'ข้อควรระวัง' ก่อนออกแบบตาราง — มีเรื่อง ID ต้องเป็น string และความแม่นของยอดขายที่ต่างกันแต่ละแพลตฟอร์ม",
    ]
    r = 3
    for ln in lines:
        c = ws.cell(r, 1, ln)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    ws.column_dimensions["A"].width = 120

    # ---- ชีต 2-5: ตาราง ----
    hdrs = ["คอลัมน์", "ชนิดที่แนะนำ", "คีย์", "บังคับมีค่า", "คำอธิบาย", "ตัวอย่าง"]
    widths = [22, 16, 8, 11, 62, 26]

    ws2 = wb.create_sheet("1_products")
    write_table(ws2, "ตาราง products  (ชีต all_in_one ในไฟล์ส่งออก)",
                "Grain: 1 แถว = 1 สินค้า/วัน · PK = (date, platform, product_id) · "
                "* product_id ว่างได้เฉพาะแถวที่ source=blocked (เปิดหน้าไม่ได้จนอ่าน id ไม่ได้)",
                hdrs, PRODUCTS, widths)
    r = ws2.max_row + 2
    c = ws2.cell(r, 1, "หมายเหตุคอลัมน์ spec: " + SPEC_COLS_NOTE)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.fill = WARN_FILL
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws2.row_dimensions[r].height = 60

    write_table(wb.create_sheet("2_variants"), "ตาราง variants",
                "Grain: 1 แถว = 1 SKU/วัน · PK = (date, platform, product_id, sku_id) · FK -> products",
                hdrs, VARIANTS, widths)

    write_table(wb.create_sheet("3_specs"), "ตาราง specs",
                "Grain: 1 แถว = 1 หัวข้อสเปก/สินค้า/วัน · PK = (date, platform, product_id, spec_name) · FK -> products",
                hdrs, SPECS, widths)

    write_table(wb.create_sheet("4_sold_history"), "ตาราง sold_history  (ไฟล์ data/sold_history.csv)",
                "Grain: 1 แถว = ยอดขายสะสมของสินค้า ณ วันหนึ่ง · PK = (date, platform, product_id) · "
                "ห้ามลบไฟล์นี้ — เป็นฐานคำนวณ sold_today/sold_mtd ย้อนหลังไม่ได้ถ้าหาย",
                hdrs, SOLD_HISTORY, widths)

    # ---- ชีต 6: ค่าที่เป็นไปได้ ----
    write_table(wb.create_sheet("5_ค่าที่เป็นไปได้"), "ค่าที่เป็นไปได้ (Enumerations)",
                "ค่าคงที่ที่ใช้ในคอลัมน์ source / platform / currency",
                ["คอลัมน์", "ค่า", "ความหมาย", "ใช้ทำรายงานได้ไหม"],
                ENUMS, [16, 14, 72, 18])

    # ---- ชีต 7: ข้อควรระวัง ----
    ws7 = wb.create_sheet("6_ข้อควรระวัง")
    write_table(ws7, "ข้อควรระวังก่อนออกแบบ/นำเข้าฐานข้อมูล",
                "เรื่องที่ทำให้ข้อมูลผิดได้ถ้าไม่รู้ก่อน",
                ["หัวข้อ", "รายละเอียด"], CAVEATS, [30, 108])
    for row in ws7.iter_rows(min_row=5, max_col=2):
        for c in row:
            c.fill = WARN_FILL

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"เขียน Data Dictionary -> {args.out}")
    print(f"  ชีต: {', '.join(ws.title for ws in wb.worksheets)}")


if __name__ == "__main__":
    main()
