#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
resolve_short.py — ตามลิงก์สั้น Lazada (s.lazada.co.th) ผ่าน HTTP redirect
เพื่อดึง URL จริง + product_id มาเติม (ลิงก์สั้นไม่มี id ในตัว)

ใช้:
    python resolve_short.py "C:\\Users\\Tada.p\\Downloads\\MarketSize-Share (1)_ids.xlsx"

- ทำเฉพาะ Lazada — TikTok (vt.tiktok) ตาม HTTP ไม่ได้ (เด้งหน้าแรก) ต้องเปิดเบราว์เซอร์ ปล่อยให้เติมตอน scrape
- ไฟล์ผลจาก fill_ids.py: คอลัมน์ D=url, E=product_id, F=shop_id, G=note (แถว 2 = ข้อมูลแถวแรก)
- เขียนทับไฟล์ *_ids.xlsx เดิม (สำรอง .bak) — ไฟล์นี้สร้างจาก openpyxl ไม่มี chart จึง round-trip ปลอดภัย
"""
import re
import sys
import time
import urllib.request
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import openpyxl

UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/150.0.0.0 Safari/537.36"


def resolve(url, timeout=20, retries=3):
    """ตาม redirect คืน (final_url, product_id) หรือ (None, เหตุผล error)
    Lazada throttle เมื่อยิงถี่ -> retry พร้อม backoff, ถือว่า 'ได้ URL แต่ไม่มี id' = ต้อง retry ด้วย"""
    last = "ไม่ได้ทำ"
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": UA})
            with urllib.request.urlopen(req, timeout=timeout) as r:
                final = r.geturl()
            m = re.search(r"-i(\d+)", final)
            if m:
                return final, m.group(1)
            last = "redirect แต่ไม่เจอ id (อาจโดน throttle เด้งหน้าอื่น)"
        except urllib.error.HTTPError as e:
            last = f"HTTP {e.code}"
            if e.code == 404:
                return None, "HTTP 404 (ลิงก์ตาย)"    # 404 = ตายจริง ไม่ต้อง retry
        except Exception as e:
            last = type(e).__name__
        time.sleep(0.5 * (attempt + 1))               # backoff
    return None, last


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    src = Path(sys.argv[1])
    if not src.exists():
        sys.exit(f"ไม่พบไฟล์: {src}")

    wb = openpyxl.load_workbook(src)
    ws = wb.active
    # เก็บแถวที่เป็นลิงก์สั้น Lazada และยังไม่มี product_id
    targets = []          # (row_index, short_url)
    for ri in range(2, ws.max_row + 1):
        url = ws.cell(row=ri, column=4).value          # D = url
        pid = ws.cell(row=ri, column=5).value          # E = product_id
        if url and "s.lazada.co.th" in str(url) and not pid:
            targets.append((ri, str(url)))

    uniq = list(dict.fromkeys(u for _, u in targets))
    print(f"ลิงก์สั้น Lazada: {len(targets)} แถว ({len(uniq)} ลิงก์ไม่ซ้ำ) — เริ่ม resolve...")

    cache = {}
    done = [0]

    def work(u):
        final, pid = resolve(u)
        cache[u] = (final, pid)
        done[0] += 1
        if done[0] % 200 == 0:
            print(f"  ...{done[0]}/{len(uniq)}", file=sys.stderr)
        time.sleep(0.05)                                # กันยิงถี่เกิน

    with ThreadPoolExecutor(max_workers=5) as ex:    # ลดเธรดกัน Lazada throttle
        list(ex.map(work, uniq))

    ok = dead = noid = 0
    for ri, u in targets:
        final, pid = cache.get(u, (None, "ไม่ได้ทำ"))
        if final and pid:
            ws.cell(row=ri, column=4).value = final     # อัปเดต url เป็นตัวจริง
            ws.cell(row=ri, column=5).value = pid        # product_id
            ws.cell(row=ri, column=7).value = ""         # เคลียร์ note
            ok += 1
        elif final and not pid:
            ws.cell(row=ri, column=7).value = "resolve แล้วแต่ไม่พบ id ใน URL"
            noid += 1
        else:
            ws.cell(row=ri, column=7).value = f"ลิงก์สั้นเสีย/ตาม redirect ไม่ได้ ({pid})"
            dead += 1

    bak = src.with_suffix(".xlsx.bak")
    if not bak.exists():
        src.replace(bak)
    else:
        bak = None
    wb.save(src)

    print(f"\nเติม product_id สำเร็จ: {ok}")
    print(f"ตาม redirect ได้แต่ไม่มี id: {noid}")
    print(f"ลิงก์เสีย/timeout: {dead}")
    if bak:
        print(f"สำรองไฟล์เดิม: {bak.name}")
    print(f"บันทึก: {src.name}")


if __name__ == "__main__":
    main()
