#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
normalize_pdp.py — แปลงผลจาก extract_pdp.js (Claude in Chrome) เป็น Excel
สคีมาเดียวกับ Master File ของ sales-tracker + คอลัมน์ spec/variation ที่ Apify ไม่มีให้

ใช้:
    python normalize_pdp.py --inputs pdp_*.json --out /mnt/user-data/outputs/PDP_2026-07-23.xlsx
    python normalize_pdp.py --inputs raw.json --date 2026-07-23 --out out.xlsx

รับได้ทั้งไฟล์ที่เป็น object เดียว, array ของ object, หรือ NDJSON (บรรทัดละ 1 object)
กติกา: ไม่เติมตัวเลขเอง ค่าที่ขาด = ว่าง + ใส่เหตุผลใน notes
"""

import argparse
import csv
import glob
import json
import os
import sys
from datetime import datetime, timedelta, timezone

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

TH = timezone(timedelta(hours=7))

PRODUCT_COLS = [
    "date", "platform", "category", "product_id", "shop_id", "shop_name", "product_name", "url",
    "price", "price_max", "original_price", "discount_pct", "currency",
    "variation_summary", "variation_count", "sku_count",
    "spec_summary", "spec_count",
    "sold_count", "sold_today", "sold_mtd", "sold_band", "source", "notes",
]
VARIANT_COLS = [
    "date", "platform", "product_id", "product_name",
    "sku_id", "option_path", "price", "original_price", "stock",
]
# spec_key = หัวข้อกลาง (canonical) จาก spec_map.csv ; spec_name = ชื่อดิบเดิมของแพลตฟอร์ม
SPEC_COLS = ["date", "platform", "product_id", "product_name", "spec_key", "spec_name", "spec_value"]


def load_spec_map(path):
    """spec_map.csv: raw_spec_name -> canonical (หัวข้อกลางข้าม 3 แพลตฟอร์ม)"""
    m = {}
    if path and os.path.exists(path):
        for row in csv.DictReader(open(path, encoding="utf-8-sig")):
            raw, canon = (row.get("raw_spec_name") or "").strip(), (row.get("canonical") or "").strip()
            if raw and canon:
                m[raw] = canon
    return m


def load_category_map(path):
    """category_map.csv: product_id -> category (จากไฟล์ MarketSize-Share)"""
    m = {}
    if path and os.path.exists(path):
        for row in csv.DictReader(open(path, encoding="utf-8-sig")):
            pid = str(row.get("product_id") or "").strip()
            if pid:
                m[pid] = (row.get("category") or "").strip()
    return m


# ---------------------------------------------------------------- loading
def load_records(paths):
    """อ่านทุกไฟล์ -> list ของ dict (รองรับ json array / object เดี่ยว / ndjson)"""
    recs = []
    for p in paths:
        with open(p, "r", encoding="utf-8") as fh:
            raw = fh.read().strip()
        if not raw:
            continue
        try:
            obj = json.loads(raw)
            recs.extend(obj if isinstance(obj, list) else [obj])
        except json.JSONDecodeError:
            for line in raw.splitlines():          # NDJSON
                line = line.strip().rstrip(",")
                if not line:
                    continue
                try:
                    recs.append(json.loads(line))
                except json.JSONDecodeError:
                    print(f"[warn] ข้ามบรรทัดที่ parse ไม่ได้ใน {p}", file=sys.stderr)
    return recs


# ---------------------------------------------------------------- helpers
def as_num(v):
    """คืน float ถ้าแปลงได้ ไม่งั้น None — ไม่เดาค่า"""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("฿", "").replace(",", "").strip()
    mult = 1.0
    if s and s[-1] in "KkMm":
        mult = 1e3 if s[-1] in "Kk" else 1e6
        s = s[:-1]
    try:
        return float(s) * mult
    except ValueError:
        return None


def fmt_variation(variation):
    """[{name, options}] -> 'สี: ดำ | เทา; ขนาด: 20V | 40V'"""
    parts = []
    for v in variation or []:
        name = (v.get("name") or "").strip()
        opts = [str(o).strip() for o in (v.get("options") or []) if str(o).strip()]
        if name:
            parts.append(f"{name}: " + " | ".join(opts) if opts else name)
    return "; ".join(parts)


def fmt_spec(spec):
    parts = []
    for s in spec or []:
        name = (s.get("name") or "").strip()
        val = s.get("value")
        if isinstance(val, list):
            val = ", ".join(str(x) for x in val)
        val = str(val).strip() if val is not None else ""
        if name and val:
            parts.append(f"{name}: {val}")
    return "; ".join(parts)


def build_rows(recs, run_date, spec_map=None, cat_map=None):
    spec_map = spec_map or {}
    cat_map = cat_map or {}
    products, variants, specs, skipped = [], [], [], []

    for r in recs:
        if not isinstance(r, dict) or r.get("error"):
            skipped.append(r.get("error") if isinstance(r, dict) else "record ไม่ใช่ object")
            continue

        pid = str(r.get("product_id") or "")           # เก็บ ID เป็น text เสมอ (TikTok 19 หลัก)
        name = r.get("product_name")
        notes = list(r.get("warnings") or [])

        price = as_num(r.get("price"))
        price_max = as_num(r.get("price_max"))
        orig = as_num(r.get("original_price"))

        if not name:
            notes.append("ไม่พบชื่อสินค้า")
        if price is None:
            notes.append("ไม่พบราคา")
        if orig is not None and price is not None and orig < price:
            notes.append("original_price < price — ตรวจสอบ")
            orig = None

        disc = round((1 - price / orig) * 100, 1) if (price and orig and orig > 0) else None

        variation = r.get("variation") or []
        spec = r.get("spec") or []
        if not variation:
            notes.append("ไม่มีข้อมูล variation")
        if not spec:
            notes.append("ไม่มีข้อมูล spec")
        if r.get("source") in ("dom", "jsonld"):
            notes.append(f"ดึงจาก {r.get('source')} (ไม่ใช่ state/api) — ความแม่นยำต่ำกว่า")

        sold_count = as_num(r.get("sold_count"))

        products.append({
            "date": run_date,
            "platform": r.get("platform"),
            "category": cat_map.get(pid, ""),      # จาก MarketSize-Share match ด้วย product_id
            "product_id": pid,
            "shop_id": str(r.get("shop_id") or ""),
            "shop_name": r.get("shop_name"),
            "product_name": name,
            "url": r.get("url"),
            "price": price,
            "price_max": price_max,
            "original_price": orig,
            "discount_pct": disc,
            "currency": r.get("currency") or "THB",
            "variation_summary": fmt_variation(variation),
            "variation_count": len(variation),
            "sku_count": len(r.get("skus") or []),
            "spec_summary": fmt_spec(spec),
            "spec_count": len(spec),
            "sold_count": sold_count,       # ยอดขายสะสม (เลขเป๊ะ)
            "sold_today": None,             # เติมทีหลังด้วย diff กับไฟล์ประวัติ
            "sold_mtd": None,               # ยอดสะสมตั้งแต่ต้นเดือน (month-to-date)
            "sold_band": r.get("sold_band"),
            "source": r.get("source"),
            "notes": "; ".join(dict.fromkeys(notes)),
        })

        for s in r.get("skus") or []:
            variants.append({
                "date": run_date,
                "platform": r.get("platform"),
                "product_id": pid,
                "product_name": name,
                "sku_id": str(s.get("sku_id") or ""),
                "option_path": s.get("option_path"),
                "price": as_num(s.get("price")),
                "original_price": as_num(s.get("original_price")),
                "stock": as_num(s.get("stock")),
            })

        for s in spec:
            val = s.get("value")
            if isinstance(val, list):
                val = ", ".join(str(x) for x in val)
            raw_name = s.get("name")
            specs.append({
                "date": run_date,
                "platform": r.get("platform"),
                "product_id": pid,
                "product_name": name,
                "spec_key": spec_map.get((raw_name or "").strip(), ""),   # ว่าง = ยังไม่ได้ map
                "spec_name": raw_name,
                "spec_value": val,
            })

    return (
        pd.DataFrame(products, columns=PRODUCT_COLS),
        pd.DataFrame(variants, columns=VARIANT_COLS),
        pd.DataFrame(specs, columns=SPEC_COLS),
        skipped,
    )


# ---------------------------------------------------------------- spec matrix
def build_spec_matrix(specs_df, prod_df):
    """pivot spec ให้เป็น 'หัวข้อเป็นคอลัมน์': 1 แถว = 1 สินค้า, คอลัมน์ = หัวข้อกลาง (canonical)
    ใช้เฉพาะ spec_key ที่ map แล้ว (ข้ามชื่อดิบที่ยังไม่ได้จับหัวข้อ) ค่าซ้ำต่อสินค้ารวมด้วย ' | '"""
    if specs_df.empty:
        return pd.DataFrame()
    df = specs_df[specs_df["spec_key"].astype(str).str.strip() != ""].copy()
    if df.empty:
        return pd.DataFrame()
    df["spec_value"] = df["spec_value"].astype(str)
    agg = (df.groupby(["platform", "product_id", "product_name", "spec_key"])["spec_value"]
             .apply(lambda s: " | ".join(dict.fromkeys(x for x in s if x.strip())))
             .reset_index())
    mat = agg.pivot_table(index=["platform", "product_id", "product_name"],
                          columns="spec_key", values="spec_value", aggfunc="first").reset_index()
    mat.columns.name = None
    if len(prod_df):
        cat = prod_df[["platform", "product_id", "category"]].drop_duplicates(["platform", "product_id"])
        mat = mat.merge(cat, on=["platform", "product_id"], how="left")
    front = [c for c in ("platform", "category", "product_id", "product_name") if c in mat.columns]
    others = sorted(c for c in mat.columns if c not in front)
    return mat[front + others]


def build_combined(prod, matrix):
    """รวมเป็นชีทเดียว 1 แถว/สินค้า: ข้อมูลสินค้า+ราคา+sold + หัวข้อ spec กลางเป็นคอลัมน์
    (= products join spec_matrix) variant ยังดูราย SKU ได้ในชีท variants"""
    if prod is None or prod.empty:
        return prod
    if matrix is None or matrix.empty:
        return prod.copy()
    topics = [c for c in matrix.columns
              if c not in ("platform", "category", "product_id", "product_name")]
    # กันแตกแถว: 1 (platform, product_id) ต้องมี matrix แถวเดียว
    m = matrix[["platform", "product_id"] + topics].drop_duplicates(["platform", "product_id"])
    return prod.merge(m, on=["platform", "product_id"], how="left")


# ---------------------------------------------------------------- sold diff
def apply_sold_diff(prod, history_path, run_date):
    """เก็บ sold_count สะสมรายวันลง history แล้วคำนวณ
       sold_today = วันนี้ - snapshot ล่าสุดก่อนหน้า
       sold_mtd   = วันนี้ - snapshot สุดท้ายก่อนวันที่ 1 ของเดือน (ยอดขายตั้งแต่ต้นเดือน)
    ID เก็บเป็น string เสมอ (TikTok 19 หลัก) กันเพี้ยน"""
    month_start = run_date[:8] + "01"        # 'YYYY-MM-01'

    hist = pd.DataFrame(columns=["date", "platform", "product_id", "sold_count"])
    if os.path.exists(history_path):
        hist = pd.read_csv(history_path, dtype={"product_id": str})
        hist = hist[hist["date"] != run_date]        # รันซ้ำวันเดิม = เขียนทับ ไม่นับซ้ำ

    def latest_before(key_hist, cutoff, inclusive=False):
        h = key_hist[key_hist["date"] <= cutoff] if inclusive else key_hist[key_hist["date"] < cutoff]
        if h.empty:
            return None
        return as_num(h.sort_values("date").iloc[-1]["sold_count"])

    def month_baseline(key_hist):
        # ยอดสะสม ณ ต้นเดือน: ใช้ snapshot ล่าสุดที่ <= วันที่ 1 (แม่นสุด)
        b = latest_before(key_hist, month_start, inclusive=True)
        if b is not None:
            return b
        # ไม่มี = เพิ่งเริ่มเก็บกลางเดือน -> ใช้ snapshot แรกของเดือนนี้ที่ก่อนวันนี้แทน
        h = key_hist[(key_hist["date"] >= month_start) & (key_hist["date"] < run_date)]
        if h.empty:
            return None
        return as_num(h.sort_values("date").iloc[0]["sold_count"])

    sold_today, sold_mtd = [], []
    for _, row in prod.iterrows():
        cur = as_num(row.get("sold_count"))
        if cur is None:
            sold_today.append(None); sold_mtd.append(None); continue
        kh = hist[(hist["platform"] == row["platform"]) & (hist["product_id"] == row["product_id"])]
        yday = latest_before(kh, run_date)                       # snapshot ล่าสุดก่อนวันนี้
        base = month_baseline(kh)                                # ยอดสะสม ณ ต้นเดือน (หรือวันแรกที่เก็บ)
        # ยอดขายต้องไม่ติดลบ (ยอดสะสมโตขึ้นเรื่อย ๆ) ถ้าติดลบแปลว่าร้านรีเซ็ต/ข้อมูลเพี้ยน -> เว้นว่าง
        sold_today.append(cur - yday if (yday is not None and cur >= yday) else None)
        sold_mtd.append(cur - base if (base is not None and cur >= base) else None)

    prod["sold_today"] = sold_today
    prod["sold_mtd"] = sold_mtd

    # เก็บ snapshot วันนี้ต่อท้าย (เฉพาะตัวที่มี sold_count)
    snap = prod.loc[prod["sold_count"].notna(), ["date", "platform", "product_id", "sold_count"]]
    out = pd.concat([hist, snap], ignore_index=True)
    os.makedirs(os.path.dirname(os.path.abspath(history_path)) or ".", exist_ok=True)
    out.to_csv(history_path, index=False)
    return prod


# ---------------------------------------------------------------- output
def fill_null(df):
    """ทุกช่องที่ไม่มีข้อมูล (NaN/None/ว่าง/ช่องว่างล้วน) -> ข้อความ "Null" ห้ามปล่อยว่าง
    ทำบนสำเนาตอนเขียน Excel เท่านั้น (ไม่กระทบ sold_history ที่คำนวณ diff)"""
    if df is None or df.empty:
        return df
    out = df.copy()
    out = out.where(out.notna(), "Null")                       # NaN/None -> Null
    out = out.replace(r"^\s*$", "Null", regex=True)            # ว่าง/ช่องว่างล้วน -> Null
    return out


def write_excel(path, dfs):
    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        for sheet, df in dfs.items():
            fill_null(df).to_excel(xw, sheet_name=sheet, index=False)

    wb = load_workbook(path)
    head_fill = PatternFill("solid", fgColor="D9D9D9")
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(name="Arial", size=10, bold=True)
            cell.fill = head_fill
            cell.alignment = Alignment(vertical="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial", size=10)
        ws.freeze_panes = "A2"
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 55)
        # บังคับคอลัมน์ ID เป็น text กัน Excel ปัดเลข 19 หลักของ TikTok
        headers = [c.value for c in ws[1]]
        for name in ("product_id", "shop_id", "sku_id"):
            if name in headers:
                letter = ws.cell(row=1, column=headers.index(name) + 1).column_letter
                for cell in ws[letter][1:]:
                    cell.number_format = "@"
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--inputs", nargs="+", required=True, help="ไฟล์ JSON จาก extract_pdp.js (ใส่ได้หลายไฟล์ / glob)")
    ap.add_argument("--date", default=None, help="YYYY-MM-DD (ค่าเริ่มต้น = วันนี้ เวลาไทย)")
    ap.add_argument("--out", required=True, help="ไฟล์ Excel ปลายทาง")
    ap.add_argument("--history", default="data/sold_history.csv",
                    help="ไฟล์เก็บ sold_count สะสมรายวัน สำหรับคำนวณ sold_today/sold_mtd")
    ap.add_argument("--spec-map", default="spec_map.csv",
                    help="ตาราง map ชื่อ spec ดิบ -> หัวข้อกลาง (สร้างด้วย build_maps.py)")
    ap.add_argument("--category-map", default="category_map.csv",
                    help="ตาราง map product_id -> category (สร้างด้วย build_maps.py)")
    args = ap.parse_args()

    paths = []
    for p in args.inputs:
        paths.extend(sorted(glob.glob(p)) or [p])
    paths = [p for p in paths if os.path.isfile(p)]
    if not paths:
        sys.exit("ไม่พบไฟล์ input")

    run_date = args.date or datetime.now(TH).strftime("%Y-%m-%d")
    spec_map = load_spec_map(args.spec_map)
    cat_map = load_category_map(args.category_map)
    recs = load_records(paths)
    prod, var, spec, skipped = build_rows(recs, run_date, spec_map, cat_map)
    if len(prod):
        prod = apply_sold_diff(prod, args.history, run_date)
    matrix = build_spec_matrix(spec, prod)
    combined = build_combined(prod, matrix)
    # ชีท all_in_one = ทุกอย่างระดับสินค้าในชีทเดียว ; variants/specs เก็บไว้ดูรายละเอียด
    write_excel(args.out, {"all_in_one": combined, "variants": var, "specs": spec})

    report = {
        "date": run_date,
        "files_read": len(paths),
        "products": len(prod),
        "by_platform": prod["platform"].value_counts().to_dict() if len(prod) else {},
        "variants": len(var),
        "spec_rows": len(spec),
        "missing_price": int(prod["price"].isna().sum()) if len(prod) else 0,
        "no_variation": int((prod["variation_count"] == 0).sum()) if len(prod) else 0,
        "no_spec": int((prod["spec_count"] == 0).sum()) if len(prod) else 0,
        "low_confidence_dom": int((prod["source"].isin(["dom", "jsonld"])).sum()) if len(prod) else 0,
        "with_sold_count": int(prod["sold_count"].notna().sum()) if len(prod) else 0,
        "with_shop_name": int(prod["shop_name"].notna().sum()) if len(prod) else 0,
        "with_category": int((prod["category"].astype(str).str.strip() != "").sum()) if len(prod) else 0,
        "spec_matrix_topics": (len([c for c in matrix.columns
                                    if c not in ("platform", "category", "product_id", "product_name")])
                               if len(matrix) else 0),
        "skipped": skipped,
        "out": args.out,
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
