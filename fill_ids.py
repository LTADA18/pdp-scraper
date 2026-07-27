#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fill_ids.py — ดึง URL จริงจาก hyperlink ที่ซ่อนในช่อง A ของ Master Raw Table
แล้วแยก product_id + shop_id ออกมา (ลิงก์ที่ copy จากมือถือเก็บ URL เป็น hyperlink แต่โชว์เป็นชื่อ)

ใช้:
    python fill_ids.py "C:\\Users\\Tada.p\\Downloads\\MarketSize-Share (1).xlsx"

ปลอดภัย: ไม่แก้ไฟล์เดิม — เขียนผลเป็นไฟล์ใหม่ *_ids.xlsx (คอลัมน์เรียงตรงกับของเดิม
คุณ copy คอลัมน์ url/product_id/shop_id ไปวางทับได้เลย เพราะลำดับแถวเหมือนเดิมเป๊ะ)
+ urls_from_master.txt (เฉพาะ URL ที่ scrape ได้ พร้อมใช้กับ scrape_pdp.py)
"""
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

SHEET = "Master Raw Table"


def platform_of(url):
    if "lazada" in url:
        return "lazada"
    if "shopee" in url:
        return "shopee"
    if "tiktok" in url:
        return "tiktok"
    return None


def parse_ids(url):
    """คืน (product_id, shop_id) จาก URL — เว้น None ตรงที่ URL ไม่มี (ไม่เดา)"""
    u = url or ""
    # Shopee: มีทั้ง shop_id + item_id — i.<shop>.<item> หรือ /product/<shop>/<item>
    m = re.search(r"i\.(\d+)\.(\d+)", u) or re.search(r"/product/(\d+)/(\d+)", u)
    if m and "shopee" in u:
        return m.group(2), m.group(1)          # (product_id, shop_id)
    # Lazada: -i<item_id>  (shop_id ไม่อยู่ใน URL)
    m = re.search(r"-i(\d+)", u)
    if m and "lazada" in u:
        return m.group(1), None
    # TikTok: /product/<id> หรือ /pdp/<slug>/<id>  (shop_id ไม่อยู่ใน URL)
    m = re.search(r"/product/(\d{6,})", u) or re.search(r"/pdp/[^/]+/(\d{6,})", u)
    if m and "tiktok" in u:
        return m.group(1), None
    return None, None


def is_shortlink(url):
    return bool(re.search(r"s\.lazada\.co\.th|vt\.tiktok\.com|/s\.[A-Za-z0-9]+$", url or ""))


def real_url(cell):
    """URL จริง = hyperlink target ถ้ามี ไม่งั้นใช้ค่าที่โชว์ถ้ามันเป็น URL อยู่แล้ว"""
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target
    v = str(cell.value or "")
    return v if v.startswith("http") else None


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    src = Path(sys.argv[1])
    if not src.exists():
        sys.exit(f"ไม่พบไฟล์: {src}")

    print(f"เปิด {src.name} (อาจช้าเพราะไฟล์ใหญ่)...")
    wb = openpyxl.load_workbook(src)              # ต้องไม่ read_only เพื่ออ่าน hyperlink
    ws = wb[SHEET]
    hdr = [c.value for c in ws[1]]
    ci = {name: i for i, name in enumerate(hdr) if name}

    out = openpyxl.Workbook()
    o = out.active
    o.title = "ids"
    # หัวตารางตรงแถว 1 เหมือนไฟล์เดิม -> แถวข้อมูลเริ่มแถว 2 ตรงกันเป๊ะ copy วางทับได้
    o.append(["orig_row", "platform", "category", "url", "product_id", "shop_id", "note"])

    urls_out = []
    n = with_url = got_pid = got_shop = short = nolink = 0

    # ออกครบทุกแถว (ไม่ข้ามแถวว่าง) เพื่อให้ลำดับตรงกับไฟล์เดิม 1:1
    for ri in range(2, ws.max_row + 1):
        a = ws.cell(row=ri, column=1)
        plat0 = ws.cell(row=ri, column=ci["platform"] + 1).value if "platform" in ci else None
        if a.value is None and not a.hyperlink and not plat0:
            o.append([ri, "", "", "", "", "", ""])   # แถวว่างจริง ใส่ว่างไว้ให้ลำดับตรง
            continue
        n += 1
        url = real_url(a)
        plat = plat0
        cat = (ws.cell(row=ri, column=ci["category"] + 1).value if "category" in ci else None)

        pid = shop = None
        note = ""
        if not url:
            note = "ไม่มี URL/hyperlink"
            nolink += 1
        else:
            with_url += 1
            if is_shortlink(url):
                note = "ลิงก์สั้น — ต้องเปิดเพื่อหา id (resolve ทีหลัง)"
                short += 1
            else:
                pid, shop = parse_ids(url)
                if pid:
                    got_pid += 1
                if shop:
                    got_shop += 1
                if not pid:
                    note = "URL ไม่มี pattern id ที่รู้จัก"
            if pid or is_shortlink(url):
                urls_out.append(url)

        o.append([ri, str(plat or ""), str(cat or ""),
                  url or "", pid or "", shop or "", note])

    dst = src.with_name(src.stem + "_ids.xlsx")
    out.save(dst)
    txt = src.with_name("urls_from_master.txt")
    txt.write_text("\n".join(dict.fromkeys(urls_out)) + "\n", encoding="utf-8")

    print(f"\nแถวข้อมูล: {n}")
    print(f"  มี URL (จาก hyperlink/ข้อความ): {with_url}")
    print(f"  แยก product_id ได้: {got_pid}")
    print(f"  แยก shop_id ได้ (ส่วนใหญ่ Shopee): {got_shop}")
    print(f"  ลิงก์สั้น (ต้อง resolve เพื่อได้ id): {short}")
    print(f"  ไม่มี URL เลย: {nolink}")
    print(f"\nไฟล์ผล: {dst.name}  (คอลัมน์เรียงตามแถวเดิม copy วางทับได้)")
    print(f"URL พร้อม scrape: {txt.name} ({len(set(urls_out))} ลิงก์)")


if __name__ == "__main__":
    main()
